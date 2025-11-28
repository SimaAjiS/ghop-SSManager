from fastapi import APIRouter, HTTPException, status
from typing import Optional, List, Dict, Any
from fastapi.responses import StreamingResponse
from sqlalchemy import (
    select,
    or_,
    asc,
    desc,
    func,
    cast,
    String,
    MetaData,
    Table,
    text,
)
import pandas as pd
import json
import os
import re
import openpyxl
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
from datetime import datetime
from ....core.config import settings
from ....core.database import get_db_engine
from ....core.utils import apply_filters, log_audit_event
from pydantic import BaseModel, Field

MAX_RELATED_NOTE_ROWS = 12

router = APIRouter()


class CharacteristicPayload(BaseModel):
    item: Optional[str] = None
    plus_minus: Optional[bool] = Field(default=None, alias="+/-")
    min: Optional[float] = None
    typ: Optional[float] = None
    max: Optional[float] = None
    unit: Optional[str] = None
    bias_vgs: Optional[str] = None
    bias_igs: Optional[str] = None
    bias_vds: Optional[str] = None
    bias_ids: Optional[str] = None
    bias_vss: Optional[str] = None
    bias_iss: Optional[str] = None
    cond: Optional[str] = None


class DeviceUpdatePayload(BaseModel):
    device: Optional[Dict[str, Any]] = Field(
        default=None, description="Columns belonging to MT_device"
    )
    spec_sheet: Optional[Dict[str, Any]] = Field(
        default=None, description="Columns belonging to MT_spec_sheet"
    )
    characteristics: Optional[List[CharacteristicPayload]] = Field(
        default=None,
        description="Complete list of MT_elec_characteristic rows. Empty list clears existing rows.",
    )


def _filter_columns(values: Optional[Dict[str, Any]], table: Table) -> Dict[str, Any]:
    if not values:
        return {}
    table_columns = {col.name for col in table.columns}
    return {k: v for k, v in values.items() if k in table_columns}


@router.get("/user/devices")
def get_user_devices(
    page: int = 1,
    limit: int = 50,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
):
    """Returns a paginated joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query with aliased columns
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        # Define column map for filters and sort
        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Count total results
        count_stmt = select(func.count()).select_from(stmt.subquery())

        # Calculate offset and apply pagination
        offset = (page - 1) * limit
        stmt = stmt.limit(limit).offset(offset)

        with engine.connect() as conn:
            # Execute count
            total_records = conn.execute(count_stmt).scalar()

            # Execute data fetch
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

        return {
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
        }

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/devices/{device_type}")
def update_device(device_type: str, payload: DeviceUpdatePayload):
    """Updates MT_device, MT_spec_sheet, and electrical characteristics for the given device."""
    try:
        engine = get_db_engine()
        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)
        mt_characteristic = Table(
            "MT_elec_characteristic", metadata, autoload_with=engine
        )

        today = datetime.utcnow().date()

        with engine.begin() as conn:
            device_row = (
                conn.execute(select(mt_device).where(mt_device.c.type == device_type))
                .mappings()
                .first()
            )
            if not device_row:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Device '{device_type}' not found",
                )

            sheet_no = device_row.get("sheet_no")
            device_changes = _filter_columns(payload.device, mt_device)
            if "type" in device_changes:
                device_changes.pop("type")
            if device_changes:
                if "更新日" in mt_device.columns and "更新日" not in device_changes:
                    device_changes["更新日"] = today

                conn.execute(
                    mt_device.update()
                    .where(mt_device.c.type == device_type)
                    .values(**device_changes)
                )

            spec_changes = _filter_columns(payload.spec_sheet, mt_spec_sheet)
            if spec_changes:
                spec_changes.pop("sheet_no", None)
                if "更新日" in mt_spec_sheet.columns and "更新日" not in spec_changes:
                    spec_changes["更新日"] = today

                conn.execute(
                    mt_spec_sheet.update()
                    .where(mt_spec_sheet.c.sheet_no == sheet_no)
                    .values(**spec_changes)
                )

            if payload.characteristics is not None:
                conn.execute(
                    mt_characteristic.delete().where(
                        mt_characteristic.c.sheet_no == sheet_no
                    )
                )
                rows_to_insert = []
                for char in payload.characteristics:
                    record = _filter_columns(
                        char.dict(by_alias=True), mt_characteristic
                    )
                    record["sheet_no"] = sheet_no
                    if "更新日" in mt_characteristic.columns and "更新日" not in record:
                        record["更新日"] = today
                    rows_to_insert.append(record)

                if rows_to_insert:
                    conn.execute(mt_characteristic.insert(), rows_to_insert)

            log_payload = {
                "device_type": device_type,
                "device_changes": device_changes,
                "spec_changes": spec_changes,
                "characteristics_count": None
                if payload.characteristics is None
                else len(payload.characteristics),
            }
            log_audit_event(
                conn,
                action="update",
                target=f"device:{device_type}",
                details=json.dumps(log_payload, ensure_ascii=False, default=str),
            )

        # Return the refreshed data for the drawer/editor
        return get_device_details(device_type)

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/user/devices/export")
def export_user_devices(
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
    format: str = "excel",
):
    """Exports joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Execute query
        with engine.connect() as conn:
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        # Convert to DataFrame
        df = pd.DataFrame(data)

        output = BytesIO()
        if format == "csv":
            df.to_csv(output, index=False)
            media_type = "text/csv"
            filename = "user_devices.csv"
        else:
            df.to_excel(output, index=False, engine="openpyxl")
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "user_devices.xlsx"

        output.seek(0)
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/devices/{device_type}/details")
def get_device_details(device_type: str):
    """Returns detailed information for a specific device, including spec sheet, maskset, and characteristics."""
    try:
        engine = get_db_engine()

        # 1. Fetch basic device info and related master data
        # We use text query for complex joins
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            # Execute device query
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )

            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )

            device_data = dict(result_device)

            # Execute elec characteristics query
            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            # Fetch all device types with the same sheet_no and their specific data
            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        COALESCE(tm.top_metal_display, d.top_metal) AS top_metal_display,
                        COALESCE(wt.wafer_thickness_display, d.wafer_thickness) AS wafer_thickness_display,
                        COALESCE(bm.back_metal_display, d.back_metal) AS back_metal_display
                    FROM MT_device d
                    LEFT JOIN MT_top_metal tm
                        ON d.top_metal = tm.top_metal
                    LEFT JOIN MT_back_metal bm
                        ON d.back_metal = bm.back_metal
                    LEFT JOIN MT_wafer_thickness wt
                        ON CAST(d.wafer_thickness AS TEXT) = CAST(wt.id AS TEXT)
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]
                if len(related_devices) > MAX_RELATED_NOTE_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"NOTE欄に出力できる関連機種は最大{MAX_RELATED_NOTE_ROWS}件です。",
                    )
                if len(related_devices) > MAX_RELATED_NOTE_ROWS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"NOTE欄に出力できる関連機種は最大{MAX_RELATED_NOTE_ROWS}件です。",
                    )

        return {
            "device": device_data,
            "characteristics": elec_data,
            "related_devices": related_devices,
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/devices/{device_type}/export-excel")
def export_device_excel(device_type: str):
    """Generates and returns an Excel spec sheet for the given device."""
    try:
        engine = get_db_engine()

        # Reuse the logic from get_device_details to fetch data
        # 1. Fetch basic device info and related master data
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )
            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )
            device_data = dict(result_device)

            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        d.top_metal as top_metal_display,
                        d.wafer_thickness as wafer_thickness_display,
                        d.back_metal as back_metal_display
                    FROM MT_device d
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]

        # Load Template
        # Updated to use the new .xlsx template
        template_path = os.path.join(
            str(settings.DATA_DIR), "templates", "specsheet_template.xlsx"
        )
        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template file not found")

        # keep_vba=True is NOT required for .xlsx files
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        if ws is None or not isinstance(ws, Worksheet):
            raise HTTPException(
                status_code=500, detail="Invalid template: no active worksheet"
            )

        # Helper to safe get
        def get_val(data, key, default=""):
            val = data.get(key)
            return val if val is not None else default

        # Helper to format condition string (ported from frontend)
        def format_condition(char):
            parts = []
            if char.get("bias_vgs"):
                parts.append(f"VGS={char.get('bias_vgs')}")
            if char.get("bias_igs"):
                parts.append(f"IGS={char.get('bias_igs')}")
            if char.get("bias_vds"):
                parts.append(f"VDS={char.get('bias_vds')}")
            if char.get("bias_ids"):
                parts.append(f"IDS={char.get('bias_ids')}")
            if char.get("bias_vss"):
                parts.append(f"VSS={char.get('bias_vss')}")
            if char.get("bias_iss"):
                parts.append(f"ISS={char.get('bias_iss')}")
            if char.get("cond"):
                parts.append(char.get("cond"))
            return ", ".join(parts)

        def format_decimal_value(value, digits=2):
            if value in (None, ""):
                return ""
            try:
                number = Decimal(str(value))
            except (InvalidOperation, ValueError, TypeError):
                return str(value)
            fmt = f"{{0:.{digits}f}}"
            return fmt.format(number)

        def format_integer_value(value, use_grouping=False):
            if value in (None, ""):
                return ""
            try:
                number = Decimal(str(value))
            except (InvalidOperation, ValueError, TypeError):
                return str(value)
            int_value = int(number.to_integral_value(rounding=ROUND_HALF_UP))
            if use_grouping:
                return f"{int_value:,}"
            return str(int_value)

        def format_limit_value(value, item):
            if value in (None, ""):
                return ""
            if item in {"IGSS", "VGSS"}:
                return f"+/-{value}"
            return value

        def format_esd_display(raw_value):
            if raw_value in (None, ""):
                return ""
            text = str(raw_value).strip()
            if not text:
                return ""

            parts = re.split(r"\s*[:：]\s*", text, maxsplit=1)
            level_text = parts[0].strip()
            descriptor = parts[1].strip().lower() if len(parts) > 1 else ""

            level_number = None
            try:
                if level_text:
                    level_number = int(
                        Decimal(level_text).to_integral_value(rounding=ROUND_HALF_UP)
                    )
            except (InvalidOperation, ValueError, TypeError):
                level_number = None

            descriptor_contains_protected = "protect" in descriptor
            descriptor_contains_non = "non" in descriptor

            if descriptor_contains_protected and descriptor_contains_non:
                return ""

            if descriptor_contains_protected:
                if not level_number or level_number <= 1:
                    return "*ESD protected"
                return f"*ESD Protected : {level_number}V"

            return text

        # Fill Data based on new template structure

        # Header Info
        ws["J5"] = get_val(device_data, "sheet_no")
        ws["N5"] = get_val(device_data, "sheet_revision")

        # Type
        ws["D7"] = get_val(device_data, "sheet_name")

        # Chip Specs
        chip_x = format_decimal_value(get_val(device_data, "chip_x_mm"), digits=2)
        chip_y = format_decimal_value(get_val(device_data, "chip_y_mm"), digits=2)
        ws["L8"] = f"{chip_x} * {chip_y} mm" if chip_x or chip_y else ""
        ws["L10"] = (
            f"{get_val(device_data, 'pad_x_gate_um')} * {get_val(device_data, 'pad_y_gate_um')} um"
        )
        ws["L11"] = (
            f"{get_val(device_data, 'pad_x_source_um')} * {get_val(device_data, 'pad_y_source_um')} um"
        )
        ws["L12"] = f"{get_val(device_data, 'dicing_line_um')} um"
        pdpw_value = format_integer_value(
            get_val(device_data, "pdpw"), use_grouping=True
        )
        ws["L16"] = f"{pdpw_value} pcs" if pdpw_value else ""

        # Chip Appearance Image (C9)
        from openpyxl.drawing.image import Image

        appearance_file = get_val(device_data, "appearance")
        image_path = None

        if appearance_file:
            potential_path = os.path.join(
                str(settings.DATA_DIR), "chip_appearances", appearance_file
            )
            if os.path.exists(potential_path):
                image_path = potential_path

        if not image_path:
            # Use placeholder
            image_path = os.path.join(
                str(settings.DATA_DIR), "chip_appearances", "no_image.png"
            )

        if os.path.exists(image_path):
            try:
                img = Image(image_path)
                # Resize image to fit 5cm (approx 189 pixels at 96 DPI)
                # 1 cm = 37.795 px
                target_size_px = 189

                # Resize keeping aspect ratio
                if img.width > 0 and img.height > 0:
                    ratio = min(target_size_px / img.width, target_size_px / img.height)
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)

                # Anchor to C9 with offset
                # Import necessary classes for advanced anchoring
                from openpyxl.drawing.spreadsheet_drawing import (
                    OneCellAnchor,
                    AnchorMarker,
                )
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.utils.units import pixels_to_EMU

                # C9 is col=2, row=8 (0-indexed)
                # Offset by 10 pixels vertically to avoid overlap with top border
                row_offset_emu = pixels_to_EMU(10)
                marker = AnchorMarker(col=2, colOff=0, row=8, rowOff=row_offset_emu)

                # Define size in EMUs
                size = XDRPositiveSize2D(
                    pixels_to_EMU(img.width), pixels_to_EMU(img.height)
                )

                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            except Exception as e:
                print(f"Failed to add image: {e}")

        # Maximum Ratings
        ws["G19"] = get_val(device_data, "vdss_V")
        ws["G20"] = get_val(device_data, "vgss_V")

        # Wafer Probing Spec (Starts at Row 25)
        # Columns: No(C), Item(D), Min(F), Typ(G), Max(H), Unit(I), Cond(J)
        start_row = 25
        base_available_rows = 10
        num_items = len(elec_data)
        extra_probe_rows = max(0, num_items - base_available_rows)
        if extra_probe_rows:
            ws.insert_rows(start_row + base_available_rows, extra_probe_rows)

        from openpyxl.styles import Alignment
        from datetime import date

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        total_probe_rows = max(base_available_rows, num_items)
        for i in range(total_probe_rows):
            row = start_row + i
            if i < num_items:
                char = elec_data[i]
                item_name = char.get("item")
                ws.cell(row=row, column=3, value=i + 1).alignment = center_align
                ws.cell(row=row, column=4, value=item_name).alignment = left_align
                ws.cell(
                    row=row,
                    column=6,
                    value=format_limit_value(char.get("min"), item_name),
                ).alignment = center_align
                ws.cell(
                    row=row,
                    column=7,
                    value=format_limit_value(char.get("typ"), item_name),
                ).alignment = center_align
                ws.cell(
                    row=row,
                    column=8,
                    value=format_limit_value(char.get("max"), item_name),
                ).alignment = center_align
                ws.cell(
                    row=row, column=9, value=char.get("unit")
                ).alignment = center_align
                ws.cell(
                    row=row, column=10, value=format_condition(char)
                ).alignment = left_align
            else:
                for col in [3, 4, 6, 7, 8, 9, 10]:
                    ws.cell(row=row, column=col, value="")

        # Tracking of row shifts for subsequent sections
        esd_base_row = 36
        related_base_row = 49
        sheet_name_base_row = 48
        update_date_base_row = 61

        # ESD row should shift only by probe insertions
        esd_row = esd_base_row + extra_probe_rows
        ws.cell(
            row=esd_row,
            column=4,
            value=format_esd_display(get_val(device_data, "esd_display")),
        )

        # Related Devices (Starts at Row 49)
        # Columns: Type(F), Top Metal(I), Wafer Thickness(K), Back Metal(M)
        related_start_row = related_base_row + extra_probe_rows
        base_related_rows = 12
        num_related = len(related_devices)
        for i in range(base_related_rows):
            row = related_start_row + i
            if i < num_related:
                dev = related_devices[i]
                ws.cell(row=row, column=6, value=dev.get("type"))
                ws.cell(row=row, column=9, value=dev.get("top_metal_display"))
                ws.cell(row=row, column=11, value=dev.get("wafer_thickness_display"))
                ws.cell(row=row, column=13, value=dev.get("back_metal_display"))
            else:
                for col in [6, 9, 11, 13]:
                    ws.cell(row=row, column=col, value="")

        # G48 (Base G48) - Sheet Name
        g48_row = sheet_name_base_row + extra_probe_rows
        ws.cell(row=g48_row, column=7, value=get_val(device_data, "sheet_name"))

        # Update Date (Base M61) shifts only with probe insertions
        update_date_row = update_date_base_row + extra_probe_rows
        ws.cell(
            row=update_date_row,
            column=13,
            value=f"'{date.today().strftime('%Y/%m/%d')}",
        ).alignment = left_align

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Filename: [sheet_no]_[sheet_name].xlsx
        s_no = get_val(device_data, "sheet_no", "X")
        s_name = get_val(device_data, "sheet_name", "X")
        if not s_no:
            s_no = "X"
        if not s_name:
            s_name = "X"

        filename = f"{s_no}_{s_name}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
