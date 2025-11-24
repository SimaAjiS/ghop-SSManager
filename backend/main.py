from fastapi import FastAPI, HTTPException
from typing import Optional
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy import create_engine, inspect, text
import os
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
import json
from io import BytesIO
from . import settings

app = FastAPI(title="Master Table Manager API")

# Mount static files for chip appearance images
CHIP_APPEARANCES_DIR = os.path.join(settings.DATA_DIR, "chip_appearances")
if os.path.exists(CHIP_APPEARANCES_DIR):
    app.mount(
        "/static/chip_appearances",
        StaticFiles(directory=CHIP_APPEARANCES_DIR),
        name="chip_appearances",
    )

# CORS configuration
origins = [
    "http://localhost:5173",  # Vite default port
    "http://127.0.0.1:5173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database configuration


def get_db_engine():
    if not os.path.exists(settings.DB_FILE):
        raise Exception("Database file not found. Please run import script first.")
    return create_engine(settings.DB_URL)


# Helper function for filtering
def apply_filters(stmt, filters_dict, column_map):
    """
    Applies filters to the SQLAlchemy statement.
    filters_dict: {col_name: filter_value}
    column_map: {col_name: sqlalchemy_column_obj}
    """
    from sqlalchemy import cast, String, and_

    conditions = []
    for col_name, value in filters_dict.items():
        if not value or col_name not in column_map:
            continue

        col = column_map[col_name]
        val_str = str(value).strip()

        # Check for operators
        # Note: This simple parsing assumes the column is numeric if using operators.
        # If casting to String for ILIKE, operators won't work as expected for numbers unless we cast back or don't cast.
        # For now, we try to parse as float for operators. If fail, fall back to string match.

        is_operator = False
        if val_str.startswith(">="):
            try:
                val = float(val_str[2:])
                conditions.append(col >= val)
                is_operator = True
            except ValueError:
                pass
        elif val_str.startswith("<="):
            try:
                val = float(val_str[2:])
                conditions.append(col <= val)
                is_operator = True
            except ValueError:
                pass
        elif val_str.startswith(">"):
            try:
                val = float(val_str[1:])
                conditions.append(col > val)
                is_operator = True
            except ValueError:
                pass
        elif val_str.startswith("<"):
            try:
                val = float(val_str[1:])
                conditions.append(col < val)
                is_operator = True
            except ValueError:
                pass

        if not is_operator:
            # Default to ILIKE for string match
            conditions.append(cast(col, String).ilike(f"%{val_str}%"))

    if conditions:
        stmt = stmt.where(and_(*conditions))

    return stmt


# Defined table order


@app.get("/api/tables")
def get_tables():
    """Returns a list of all tables in the database."""
    try:
        engine = get_db_engine()
        inspector = inspect(engine)
        tables = inspector.get_table_names()

        # Sort tables based on TABLE_ORDER
        # Tables not in TABLE_ORDER will be appended at the end, sorted alphabetically
        def sort_key(table_name):
            if table_name in settings.TABLE_ORDER:
                return settings.TABLE_ORDER.index(table_name)
            return len(settings.TABLE_ORDER) + 1  # Put at the end

        # Separate tables into those in the list and those not
        ordered_tables = [t for t in settings.TABLE_ORDER if t in tables]
        other_tables = sorted([t for t in tables if t not in settings.TABLE_ORDER])

        final_tables = ordered_tables + other_tables

        return {"tables": final_tables}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tables/{table_name}")
def get_table_data(
    table_name: str,
    page: int = 1,
    limit: int = 50,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
):
    """Returns paginated data for a specific table with optional search, sort, and column filters."""
    try:
        engine = get_db_engine()
        inspector = inspect(engine)
        if table_name not in inspector.get_table_names():
            raise HTTPException(
                status_code=404, detail=f"Table '{table_name}' not found"
            )

        # Calculate offset
        offset = (page - 1) * limit

        # Build query
        from sqlalchemy import (
            MetaData,
            Table,
            select,
            or_,
            asc,
            desc,
            func,
            cast,
            String,
        )

        metadata = MetaData()
        table = Table(table_name, metadata, autoload_with=engine)

        # Base query
        stmt = select(table)

        # Apply Global Search
        if search:
            search_conditions = []
            for column in table.columns:
                search_conditions.append(cast(column, String).ilike(f"%{search}%"))

            if search_conditions:
                stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                # Map column names to column objects
                column_map = {c.name: c for c in table.columns}
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass  # Ignore invalid JSON

        # Apply Sort
        if sort_by:
            if sort_by in table.columns:
                col = table.columns[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Count total results (before pagination)
        count_stmt = select(func.count()).select_from(stmt.subquery())

        # Apply Pagination
        stmt = stmt.limit(limit).offset(offset)

        with engine.connect() as conn:
            # Execute count
            total_records = conn.execute(count_stmt).scalar()

            # Execute data fetch
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

        return {
            "table": table_name,
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/tables/{table_name}/export")
def export_table_data(
    table_name: str,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
    format: str = "excel",  # excel or csv
):
    """Exports data for a specific table with optional search, sort, and column filters."""
    try:
        engine = get_db_engine()
        inspector = inspect(engine)
        if table_name not in inspector.get_table_names():
            raise HTTPException(
                status_code=404, detail=f"Table '{table_name}' not found"
            )

        from sqlalchemy import (
            MetaData,
            Table,
            select,
            or_,
            asc,
            desc,
            cast,
            String,
        )

        metadata = MetaData()
        table = Table(table_name, metadata, autoload_with=engine)

        # Base query
        stmt = select(table)

        # Apply Global Search
        if search:
            search_conditions = []
            for column in table.columns:
                search_conditions.append(cast(column, String).ilike(f"%{search}%"))

            if search_conditions:
                stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                column_map = {c.name: c for c in table.columns}
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in table.columns:
                col = table.columns[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Execute query
        with engine.connect() as conn:
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        # Convert to DataFrame
        df = pd.DataFrame(data)

        output = BytesIO()
        if format == "csv":
            df.to_csv(output, index=False)
            media_type = "text/csv"
            filename = f"{table_name}.csv"
        else:
            df.to_excel(output, index=False, engine="openpyxl")
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"{table_name}.xlsx"

        output.seek(0)
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/user/devices")
def get_user_devices(
    page: int = 1,
    limit: int = 50,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
):
    """Returns a paginated joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        from sqlalchemy import (
            MetaData,
            Table,
            select,
            or_,
            asc,
            desc,
            func,
            cast,
            String,
        )

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query with aliased columns
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        # Define column map for filters and sort
        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Count total results
        count_stmt = select(func.count()).select_from(stmt.subquery())

        # Calculate offset and apply pagination
        offset = (page - 1) * limit
        stmt = stmt.limit(limit).offset(offset)

        with engine.connect() as conn:
            # Execute count
            total_records = conn.execute(count_stmt).scalar()

            # Execute data fetch
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

        return {
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
        }

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/user/devices/export")
def export_user_devices(
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
    format: str = "excel",
):
    """Exports joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        from sqlalchemy import (
            MetaData,
            Table,
            select,
            or_,
            asc,
            desc,
            cast,
            String,
        )

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Execute query
        with engine.connect() as conn:
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        # Convert to DataFrame
        df = pd.DataFrame(data)

        output = BytesIO()
        if format == "csv":
            df.to_csv(output, index=False)
            media_type = "text/csv"
            filename = "user_devices.csv"
        else:
            df.to_excel(output, index=False, engine="openpyxl")
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "user_devices.xlsx"

        output.seek(0)
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/devices/{device_type}/details")
def get_device_details(device_type: str):
    """Returns detailed information for a specific device, including spec sheet, maskset, and characteristics."""
    try:
        engine = get_db_engine()

        # 1. Fetch basic device info and related master data
        # We use text query for complex joins
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            # Execute device query
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )

            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )

            device_data = dict(result_device)

            # Execute elec characteristics query
            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            # Fetch all device types with the same sheet_no and their specific data
            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        d.top_metal as top_metal_display,
                        d.wafer_thickness as wafer_thickness_display,
                        d.back_metal as back_metal_display
                    FROM MT_device d
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]

        return {
            "device": device_data,
            "characteristics": elec_data,
            "related_devices": related_devices,
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/devices/{device_type}/export-excel")
def export_device_excel(device_type: str):
    """Generates and returns an Excel spec sheet for the given device."""
    try:
        engine = get_db_engine()

        # Reuse the logic from get_device_details to fetch data
        # 1. Fetch basic device info and related master data
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )
            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )
            device_data = dict(result_device)

            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        d.top_metal as top_metal_display,
                        d.wafer_thickness as wafer_thickness_display,
                        d.back_metal as back_metal_display
                    FROM MT_device d
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]

        # Load Template
        # Updated to use the new .xlsm template
        template_path = os.path.join(
            settings.DATA_DIR, "templates", "specsheet_template.xlsm"
        )
        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template file not found")

        # keep_vba=True is required for .xlsm files
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        if ws is None or not isinstance(ws, Worksheet):
            raise HTTPException(
                status_code=500, detail="Invalid template: no active worksheet"
            )

        # Helper to safe get
        def get_val(data, key, default=""):
            val = data.get(key)
            return val if val is not None else default

        # Fill Data based on new template structure

        # Header Info
        ws["K3"] = get_val(device_data, "sheet_no")
        ws["O3"] = get_val(device_data, "sheet_revision")

        # Type
        ws["D7"] = get_val(device_data, "sheet_name")

        # Chip Specs
        ws["H8"] = (
            f"{get_val(device_data, 'chip_x_mm')} * {get_val(device_data, 'chip_y_mm')} mm"
        )
        ws["L10"] = (
            f"{get_val(device_data, 'pad_x_gate_um')} * {get_val(device_data, 'pad_y_gate_um')} um"
        )
        ws["L11"] = (
            f"{get_val(device_data, 'pad_x_source_um')} * {get_val(device_data, 'pad_y_source_um')} um"
        )
        ws["H12"] = f"{get_val(device_data, 'dicing_line_um')} um"
        ws["H16"] = f"{get_val(device_data, 'pdpw')}pcs"

        # Maximum Ratings
        ws["G19"] = get_val(device_data, "vdss_V")
        ws["G20"] = get_val(device_data, "vgss_V")

        # Wafer Probing Spec (Starts at Row 25)
        # Columns: No(C), Item(D), Min(F), Typ(G), Max(H), Unit(I), Cond(J)
        start_row = 25
        available_rows = 10

        from openpyxl.styles import Alignment, Border, Side

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Insert rows for characteristics
        if elec_data:
            num_items = len(elec_data)

            # Only insert if we have more items than available rows
            if num_items > available_rows:
                ws.insert_rows(
                    start_row + available_rows, amount=num_items - available_rows
                )

            # Iterate over max(num_items, available_rows) to ensure we fill data AND clear unused rows
            for i in range(max(num_items, available_rows)):
                row = start_row + i
                if i < num_items:
                    char = elec_data[i]
                    ws.cell(
                        row=row, column=3, value=i + 1
                    ).alignment = center_align  # No (C)
                    ws.cell(row=row, column=3).border = thin_border

                    ws.cell(
                        row=row, column=4, value=char.get("item")
                    ).alignment = center_align  # Item (D)
                    ws.cell(row=row, column=4).border = thin_border

                    ws.cell(
                        row=row, column=6, value=char.get("min")
                    ).alignment = center_align  # Min (F)
                    ws.cell(row=row, column=6).border = thin_border

                    ws.cell(
                        row=row, column=7, value=char.get("typ")
                    ).alignment = center_align  # Typ (G)
                    ws.cell(row=row, column=7).border = thin_border

                    ws.cell(
                        row=row, column=8, value=char.get("max")
                    ).alignment = center_align  # Max (H)
                    ws.cell(row=row, column=8).border = thin_border

                    ws.cell(
                        row=row, column=9, value=char.get("unit")
                    ).alignment = center_align  # Unit (I)
                    ws.cell(row=row, column=9).border = thin_border

                    ws.cell(
                        row=row, column=10, value=char.get("cond")
                    ).alignment = center_align  # Cond (J)
                    ws.cell(row=row, column=10).border = thin_border
                else:
                    # Clear unused rows (content only, keep formatting/borders if desired, or clear all)
                    # User requested to avoid insertion, implying we should use the blank rows.
                    # Clearing content makes them blank.
                    for col in [3, 4, 6, 7, 8, 9, 10]:
                        ws.cell(row=row, column=col, value="")

        # Notes Section
        # NOTE: is at Row 47.
        # Shift depends on how many rows we INSERTED.
        # Inserted = max(0, num_items - available_rows)
        shift_amount = max(0, len(elec_data) - available_rows) if elec_data else 0

        # ESD Display
        if get_val(device_data, "esd_display"):
            ws[f"C{47 + shift_amount}"] = (
                f"NOTE: *{get_val(device_data, 'esd_display')}"
            )

        # Footer Table (Related Devices)
        # Header at Row 48 + shift
        # Data starts at Row 49 + shift
        footer_header_row = 48 + shift_amount
        footer_start_row = 49 + shift_amount
        footer_available_rows = 10  # Assumed based on analysis

        # Update Footer Header with Sheet Name
        ws[f"F{footer_header_row}"] = f"TYPE: {get_val(device_data, 'sheet_name')}"

        if related_devices:
            num_rel = len(related_devices)

            if num_rel > footer_available_rows:
                ws.insert_rows(
                    footer_start_row + footer_available_rows,
                    amount=num_rel - footer_available_rows,
                )

            for i in range(max(num_rel, footer_available_rows)):
                row = footer_start_row + i
                if i < num_rel:
                    dev = related_devices[i]

                    # Type (F)
                    ws.cell(
                        row=row, column=6, value=dev.get("type")
                    ).alignment = center_align
                    ws.cell(row=row, column=6).border = thin_border

                    # Top Metal (I)
                    ws.cell(
                        row=row, column=9, value=dev.get("top_metal_display")
                    ).alignment = center_align
                    ws.cell(row=row, column=9).border = thin_border

                    # Wafer Thickness (K)
                    ws.cell(
                        row=row, column=11, value=dev.get("wafer_thickness_display")
                    ).alignment = center_align
                    ws.cell(row=row, column=11).border = thin_border

                    # Back Metal (M)
                    ws.cell(
                        row=row, column=13, value=dev.get("back_metal_display")
                    ).alignment = center_align
                    ws.cell(row=row, column=13).border = thin_border
                else:
                    # Clear unused rows
                    for col in [6, 9, 11, 13]:
                        ws.cell(row=row, column=col, value="")

        # Save to a temporary buffer
        from io import BytesIO

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{get_val(device_data, 'sheet_no')}_{get_val(device_data, 'sheet_name')}.xlsm"

        from fastapi.responses import StreamingResponse

        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(
            output,
            headers=headers,
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        )

    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/audit-logs")
def get_audit_logs(
    page: int = 1,
    limit: int = 50,
    sort_by: Optional[str] = "timestamp",
    descending: bool = True,
):
    """Returns paginated audit logs."""
    try:
        engine = get_db_engine()

        from sqlalchemy import MetaData, Table, select, desc, asc, func

        metadata = MetaData()
        audit_log = Table("AuditLog", metadata, autoload_with=engine)

        stmt = select(audit_log)

        # Apply Sort
        if sort_by and sort_by in audit_log.columns:
            col = audit_log.columns[sort_by]
            if descending:
                stmt = stmt.order_by(desc(col))
            else:
                stmt = stmt.order_by(asc(col))
        else:
            # Default sort
            stmt = stmt.order_by(desc(audit_log.c.timestamp))

        # Count total
        count_stmt = select(func.count()).select_from(stmt.subquery())

        # Pagination
        offset = (page - 1) * limit
        stmt = stmt.limit(limit).offset(offset)

        with engine.connect() as conn:
            total_records = conn.execute(count_stmt).scalar()
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

        return {
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
        }

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
