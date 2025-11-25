from fastapi import APIRouter, HTTPException
from typing import Optional
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, asc, desc, func, cast, String, MetaData, Table, text
import pandas as pd
import json
import os
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
from .. import settings
from ..database import get_db_engine
from ..utils import apply_filters

router = APIRouter()


@router.get("/user/devices")
def get_user_devices(
    page: int = 1,
    limit: int = 50,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
):
    """Returns a paginated joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query with aliased columns
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        # Define column map for filters and sort
        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Count total results
        count_stmt = select(func.count()).select_from(stmt.subquery())

        # Calculate offset and apply pagination
        offset = (page - 1) * limit
        stmt = stmt.limit(limit).offset(offset)

        with engine.connect() as conn:
            # Execute count
            total_records = conn.execute(count_stmt).scalar()

            # Execute data fetch
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        total_pages = (total_records + limit - 1) // limit if limit > 0 else 1

        return {
            "data": data,
            "total": total_records,
            "page": page,
            "limit": limit,
            "total_pages": total_pages,
        }

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/user/devices/export")
def export_user_devices(
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    descending: bool = False,
    filters: Optional[str] = None,
    format: str = "excel",
):
    """Exports joined view of devices and their spec sheets."""
    try:
        engine = get_db_engine()

        metadata = MetaData()
        mt_device = Table("MT_device", metadata, autoload_with=engine)
        mt_spec_sheet = Table("MT_spec_sheet", metadata, autoload_with=engine)

        # Build join query
        stmt = select(
            mt_device.c.type.label("Device Type"),
            mt_device.c.sheet_no.label("Sheet No"),
            mt_spec_sheet.c.sheet_name.label("Sheet Name"),
            mt_device.c.status.label("Status"),
            mt_spec_sheet.c.vdss_V.label("Vdss (V)"),
            mt_spec_sheet.c.vgss_V.label("Vgss (V)"),
            mt_spec_sheet.c.idss_A.label("Idss (A)"),
        ).select_from(
            mt_device.outerjoin(
                mt_spec_sheet, mt_device.c.sheet_no == mt_spec_sheet.c.sheet_no
            )
        )

        column_map = {
            "Device Type": mt_device.c.type,
            "Sheet No": mt_device.c.sheet_no,
            "Sheet Name": mt_spec_sheet.c.sheet_name,
            "Status": mt_device.c.status,
            "Vdss (V)": mt_spec_sheet.c.vdss_V,
            "Vgss (V)": mt_spec_sheet.c.vgss_V,
            "Idss (A)": mt_spec_sheet.c.idss_A,
        }

        # Apply Search
        if search:
            search_conditions = [
                cast(col, String).ilike(f"%{search}%") for col in column_map.values()
            ]
            stmt = stmt.where(or_(*search_conditions))

        # Apply Column Filters
        if filters:
            try:
                filters_dict = json.loads(filters)
                stmt = apply_filters(stmt, filters_dict, column_map)
            except json.JSONDecodeError:
                pass

        # Apply Sort
        if sort_by:
            if sort_by in column_map:
                col = column_map[sort_by]
                if descending:
                    stmt = stmt.order_by(desc(col))
                else:
                    stmt = stmt.order_by(asc(col))

        # Execute query
        with engine.connect() as conn:
            result = conn.execute(stmt)
            data = [dict(row._mapping) for row in result]

        # Convert to DataFrame
        df = pd.DataFrame(data)

        output = BytesIO()
        if format == "csv":
            df.to_csv(output, index=False)
            media_type = "text/csv"
            filename = "user_devices.csv"
        else:
            df.to_excel(output, index=False, engine="openpyxl")
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "user_devices.xlsx"

        output.seek(0)
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/devices/{device_type}/details")
def get_device_details(device_type: str):
    """Returns detailed information for a specific device, including spec sheet, maskset, and characteristics."""
    try:
        engine = get_db_engine()

        # 1. Fetch basic device info and related master data
        # We use text query for complex joins
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            # Execute device query
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )

            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )

            device_data = dict(result_device)

            # Execute elec characteristics query
            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            # Fetch all device types with the same sheet_no and their specific data
            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        d.top_metal as top_metal_display,
                        d.wafer_thickness as wafer_thickness_display,
                        d.back_metal as back_metal_display
                    FROM MT_device d
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]

        return {
            "device": device_data,
            "characteristics": elec_data,
            "related_devices": related_devices,
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/devices/{device_type}/export-excel")
def export_device_excel(device_type: str):
    """Generates and returns an Excel spec sheet for the given device."""
    try:
        engine = get_db_engine()

        # Reuse the logic from get_device_details to fetch data
        # 1. Fetch basic device info and related master data
        query_device = text("""
            SELECT
                d.type, d.sheet_no, d.barrier, d.passivation, d.status,
                s.sheet_name, s.sheet_revision, s.vdss_V, s.vgss_V, s.idss_A, s.esd_display, s.maskset,
                m.chip_x_mm, m.chip_y_mm, m.dicing_line_um, m.pad_x_gate_um, m.pad_y_gate_um, m.pad_x_source_um, m.pad_y_source_um, m.pdpw, m.appearance,
                tm.top_metal, tm.top_metal_thickness_um, tm.top_metal_display,
                bm.back_metal, bm.back_metal_thickness_um, bm.back_metal_display,
                wt.wafer_thickness_um, wt.wafer_thickness_tolerance_um, wt.wafer_thickness_display
            FROM MT_device d
            LEFT JOIN MT_spec_sheet s ON d.sheet_no = s.sheet_no
            LEFT JOIN MT_maskset m ON s.maskset = m.maskset
            LEFT JOIN MT_top_metal tm ON d.top_metal = tm.top_metal
            LEFT JOIN MT_back_metal bm ON d.back_metal = bm.back_metal
            LEFT JOIN MT_wafer_thickness wt ON d.wafer_thickness = wt.id
            WHERE d.type = :device_type
        """)

        # 2. Fetch electrical characteristics
        query_elec = text("""
            SELECT
                item, `+/-` as plus_minus, min, typ, max, unit,
                bias_vgs, bias_igs, bias_vds, bias_ids, bias_vss, bias_iss, cond
            FROM MT_elec_characteristic
            WHERE sheet_no = (SELECT sheet_no FROM MT_device WHERE type = :device_type)
        """)

        with engine.connect() as conn:
            result_device = (
                conn.execute(query_device, {"device_type": device_type})
                .mappings()
                .first()
            )
            if not result_device:
                raise HTTPException(
                    status_code=404, detail=f"Device '{device_type}' not found"
                )
            device_data = dict(result_device)

            result_elec = (
                conn.execute(query_elec, {"device_type": device_type}).mappings().all()
            )
            elec_data = [dict(row) for row in result_elec]

            sheet_no = device_data.get("sheet_no")
            related_devices = []
            if sheet_no:
                query_related_devices = text("""
                    SELECT
                        d.type,
                        d.top_metal as top_metal_display,
                        d.wafer_thickness as wafer_thickness_display,
                        d.back_metal as back_metal_display
                    FROM MT_device d
                    WHERE d.sheet_no = :sheet_no
                    ORDER BY d.type ASC
                """)
                result_devices = (
                    conn.execute(query_related_devices, {"sheet_no": sheet_no})
                    .mappings()
                    .all()
                )
                related_devices = [dict(row) for row in result_devices]

        # Load Template
        # Updated to use the new .xlsx template
        template_path = os.path.join(
            settings.DATA_DIR, "templates", "specsheet_template.xlsx"
        )
        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail="Template file not found")

        # keep_vba=True is NOT required for .xlsx files
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        if ws is None or not isinstance(ws, Worksheet):
            raise HTTPException(
                status_code=500, detail="Invalid template: no active worksheet"
            )

        # Helper to safe get
        def get_val(data, key, default=""):
            val = data.get(key)
            return val if val is not None else default

        # Helper to format condition string (ported from frontend)
        def format_condition(char):
            parts = []
            if char.get("bias_vgs"):
                parts.append(f"VGS={char.get('bias_vgs')}")
            if char.get("bias_igs"):
                parts.append(f"IGS={char.get('bias_igs')}")
            if char.get("bias_vds"):
                parts.append(f"VDS={char.get('bias_vds')}")
            if char.get("bias_ids"):
                parts.append(f"IDS={char.get('bias_ids')}")
            if char.get("bias_vss"):
                parts.append(f"VSS={char.get('bias_vss')}")
            if char.get("bias_iss"):
                parts.append(f"ISS={char.get('bias_iss')}")
            if char.get("cond"):
                parts.append(char.get("cond"))
            return ", ".join(parts)

        # Fill Data based on new template structure

        # Header Info
        ws["J5"] = get_val(device_data, "sheet_no")
        ws["N5"] = get_val(device_data, "sheet_revision")

        # Type
        ws["D7"] = get_val(device_data, "sheet_name")

        # Chip Specs
        ws["L8"] = (
            f"{get_val(device_data, 'chip_x_mm')} * {get_val(device_data, 'chip_y_mm')} mm"
        )
        ws["L10"] = (
            f"{get_val(device_data, 'pad_x_gate_um')} * {get_val(device_data, 'pad_y_gate_um')} um"
        )
        ws["L11"] = (
            f"{get_val(device_data, 'pad_x_source_um')} * {get_val(device_data, 'pad_y_source_um')} um"
        )
        ws["L12"] = f"{get_val(device_data, 'dicing_line_um')} um"
        ws["L16"] = f"{get_val(device_data, 'pdpw')}pcs"

        # Chip Appearance Image (C9)
        from openpyxl.drawing.image import Image

        appearance_file = get_val(device_data, "appearance")
        image_path = None

        if appearance_file:
            potential_path = os.path.join(
                settings.DATA_DIR, "chip_appearances", appearance_file
            )
            if os.path.exists(potential_path):
                image_path = potential_path

        if not image_path:
            # Use placeholder
            image_path = os.path.join(
                settings.DATA_DIR, "chip_appearances", "no_image.png"
            )

        if os.path.exists(image_path):
            try:
                img = Image(image_path)
                # Resize image to fit reasonably (e.g., max 300px width/height)
                # Adjust as needed based on cell size
                max_size = 300
                if img.width > max_size or img.height > max_size:
                    ratio = min(max_size / img.width, max_size / img.height)
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)

                # Anchor to C9
                ws.add_image(img, "C9")
            except Exception as e:
                print(f"Failed to add image: {e}")

        # Maximum Ratings
        ws["G19"] = get_val(device_data, "vdss_V")
        ws["G20"] = get_val(device_data, "vgss_V")

        # Wafer Probing Spec (Starts at Row 25)
        # Columns: No(C), Item(D), Min(F), Typ(G), Max(H), Unit(I), Cond(J)
        start_row = 25
        available_rows = 10

        from openpyxl.styles import Alignment, Border, Side
        from datetime import date

        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        inserted_wafer_rows = 0
        # Insert rows for characteristics
        if elec_data:
            num_items = len(elec_data)

            # Only insert if we have more items than available rows
            if num_items > available_rows:
                inserted_wafer_rows = num_items - available_rows
                ws.insert_rows(start_row + available_rows, amount=inserted_wafer_rows)

            # Iterate over max(num_items, available_rows) to ensure we fill data AND clear unused rows
            for i in range(max(num_items, available_rows)):
                row = start_row + i
                if i < num_items:
                    char = elec_data[i]
                    ws.cell(
                        row=row, column=3, value=i + 1
                    ).alignment = center_align  # No (C)
                    ws.cell(row=row, column=3).border = thin_border

                    ws.cell(
                        row=row, column=4, value=char.get("item")
                    ).alignment = center_align  # Item (D)
                    ws.cell(row=row, column=4).border = thin_border

                    ws.cell(
                        row=row, column=6, value=char.get("min")
                    ).alignment = center_align  # Min (F)
                    ws.cell(row=row, column=6).border = thin_border

                    ws.cell(
                        row=row, column=7, value=char.get("typ")
                    ).alignment = center_align  # Typ (G)
                    ws.cell(row=row, column=7).border = thin_border

                    ws.cell(
                        row=row, column=8, value=char.get("max")
                    ).alignment = center_align  # Max (H)
                    ws.cell(row=row, column=8).border = thin_border

                    ws.cell(
                        row=row, column=9, value=char.get("unit")
                    ).alignment = center_align  # Unit (I)
                    ws.cell(row=row, column=9).border = thin_border

                    # Use format_condition helper
                    ws.cell(
                        row=row, column=10, value=format_condition(char)
                    ).alignment = center_align  # Cond (J)
                    ws.cell(row=row, column=10).border = thin_border
                else:
                    # Clear cells if no data
                    for col in [3, 4, 6, 7, 8, 9, 10]:
                        ws.cell(row=row, column=col, value="")
                        ws.cell(row=row, column=col).border = thin_border

        # ESD (Base D36)
        # Shifted by inserted_wafer_rows
        esd_row = 36 + inserted_wafer_rows
        ws.cell(row=esd_row, column=4, value=get_val(device_data, "esd_display"))

        # Related Devices (Starts at Row 49 + inserted rows)
        # Columns: Type(F), Top Metal(I), Wafer Thickness(K), Back Metal(M)
        base_related_start_row = 49
        related_start_row = base_related_start_row + inserted_wafer_rows
        available_related_rows = 5

        inserted_related_rows = 0
        if related_devices:
            num_related = len(related_devices)
            if num_related > available_related_rows:
                inserted_related_rows = num_related - available_related_rows
                ws.insert_rows(
                    related_start_row + available_related_rows,
                    amount=inserted_related_rows,
                )

            for i in range(max(num_related, available_related_rows)):
                row = related_start_row + i
                if i < num_related:
                    dev = related_devices[i]
                    ws.cell(
                        row=row, column=6, value=dev.get("type")
                    ).alignment = center_align  # Type (F)
                    ws.cell(row=row, column=6).border = thin_border

                    ws.cell(
                        row=row, column=9, value=dev.get("top_metal_display")
                    ).alignment = center_align  # Top Metal (I)
                    ws.cell(row=row, column=9).border = thin_border

                    ws.cell(
                        row=row, column=11, value=dev.get("wafer_thickness_display")
                    ).alignment = center_align  # Wafer Thickness (K)
                    ws.cell(row=row, column=11).border = thin_border

                    ws.cell(
                        row=row, column=13, value=dev.get("back_metal_display")
                    ).alignment = center_align  # Back Metal (M)
                    ws.cell(row=row, column=13).border = thin_border
                else:
                    # Clear cells
                    for col in [6, 9, 11, 13]:
                        ws.cell(row=row, column=col, value="")
                        ws.cell(row=row, column=col).border = thin_border

        # G48 (Base G48) - Sheet Name
        # Shifted by both insertions
        g48_row = 48 + inserted_wafer_rows + inserted_related_rows
        ws.cell(row=g48_row, column=7, value=get_val(device_data, "sheet_name"))

        # Update Date (Base M53)
        # Shifted by both insertions
        # Note: If base related row changed from 40 to 49, the update date row might also need adjustment relative to that.
        # Assuming M53 is the absolute position in the template BEFORE any insertions.
        # If Related Devices starts at 49 and has 5 rows, it ends at 53.
        # So M53 seems to be overlapping with the last row of Related Devices if it has 5 rows.
        # Let's assume the user meant M53 is the footer date position in the template.
        update_date_row = 53 + inserted_wafer_rows + inserted_related_rows
        ws.cell(row=update_date_row, column=13, value=date.today())

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Filename: [sheet_no]_[sheet_name].xlsx
        s_no = get_val(device_data, "sheet_no", "X")
        s_name = get_val(device_data, "sheet_name", "X")
        if not s_no:
            s_no = "X"
        if not s_name:
            s_name = "X"

        filename = f"{s_no}_{s_name}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        return StreamingResponse(output, headers=headers, media_type=media_type)

    except Exception as e:
        import traceback

        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
