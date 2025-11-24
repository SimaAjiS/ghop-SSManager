import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spec Sheet"

    # Set column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
    )

    # Header
    ws.merge_cells("A1:G1")
    ws["A1"] = "SPEC SHEET (FOR REFERENCE)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Sheet Info
    ws["E2"] = "SHEET No."
    ws["F2"] = "{{sheet_no}}"
    ws["E3"] = "Rev."
    ws["F3"] = "{{sheet_revision}}"
    ws["E4"] = "Page."
    ws["F4"] = "1 of 1"

    for row in range(2, 5):
        ws[f"E{row}"].font = bold_font
        ws[f"E{row}"].alignment = center_align
        ws[f"F{row}"].alignment = center_align
        ws[f"E{row}"].border = thin_border
        ws[f"F{row}"].border = thin_border

    # Type
    ws.merge_cells("A6:C6")
    ws["A6"] = "TYPE: {{type}}"
    ws["A6"].font = Font(bold=True, size=12)

    # Chip Specs Section
    ws.merge_cells("A8:G8")
    ws["A8"] = "CHIP SPECS"
    ws["A8"].font = bold_font
    ws["A8"].fill = header_fill
    ws["A8"].border = thin_border

    specs = [
        ("CHIP SIZE", "{{chip_size}}"),
        ("CHIP THICKNESS", 'Refer to "NOTE"'),
        ("BONDING PAD DIMENSIONS", ""),
        ("  GATE", "{{pad_gate}}"),
        ("  SOURCE", "{{pad_source}}"),
        ("SCRIBE LINE WIDTH", "{{dicing_line_um}} um"),
        ("TOP METAL", 'Refer to "NOTE"'),
        ("BACK METAL", 'Refer to "NOTE"'),
        ("WAFER SIZE", "6inch"),
        ("POSSIBLE DIE PER WAFER", "{{pdpw}}pcs"),
    ]

    row_idx = 9
    for label, value in specs:
        ws.merge_cells(f"A{row_idx}:C{row_idx}")
        ws[f"A{row_idx}"] = label
        ws.merge_cells(f"D{row_idx}:G{row_idx}")
        ws[f"D{row_idx}"] = value

        ws[f"A{row_idx}"].border = thin_border
        ws[f"D{row_idx}"].border = thin_border
        ws[f"A{row_idx}"].alignment = left_align
        ws[f"D{row_idx}"].alignment = left_align

        row_idx += 1

    # Maximum Ratings
    row_idx += 2
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    ws[f"A{row_idx}"] = "Maximum Ratings(Ta=25C) (FOR REFERENCE)"
    ws[f"A{row_idx}"].font = bold_font
    ws[f"A{row_idx}"].fill = header_fill
    ws[f"A{row_idx}"].border = thin_border

    row_idx += 1
    headers = ["Characteristics", "Symbol", "Ratings", "Unit"]

    ws.merge_cells(f"A{row_idx}:C{row_idx}")  # Characteristics
    ws[f"A{row_idx}"] = headers[0]
    ws[f"D{row_idx}"] = headers[1]
    ws.merge_cells(f"E{row_idx}:F{row_idx}")  # Ratings
    ws[f"E{row_idx}"] = headers[2]
    ws[f"G{row_idx}"] = headers[3]

    for col in ["A", "D", "E", "G"]:
        cell = ws[f"{col}{row_idx}"]
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    row_idx += 1
    ratings = [
        ("Drain-source voltage", "VDSS", "{{vdss_V}}", "V"),
        ("Gate-source voltage", "VGSS", "{{vgss_V}}", "V"),
    ]

    for char, sym, rate, unit in ratings:
        ws.merge_cells(f"A{row_idx}:C{row_idx}")
        ws[f"A{row_idx}"] = char
        ws[f"D{row_idx}"] = sym
        ws.merge_cells(f"E{row_idx}:F{row_idx}")
        ws[f"E{row_idx}"] = rate
        ws[f"G{row_idx}"] = unit

        for col in ["A", "D", "E", "G"]:
            ws[f"{col}{row_idx}"].border = thin_border
            ws[f"{col}{row_idx}"].alignment = center_align
        ws[f"A{row_idx}"].alignment = left_align

        row_idx += 1

    # Wafer Probing Spec
    row_idx += 2
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    ws[f"A{row_idx}"] = "WAFER PROBING SPEC (Ta=25C)"
    ws[f"A{row_idx}"].font = bold_font
    ws[f"A{row_idx}"].fill = header_fill
    ws[f"A{row_idx}"].border = thin_border

    row_idx += 1
    # Table Header
    # No | MODE | LIMIT (MIN, Typ, MAX) | UNIT | CONDITIONS
    ws.merge_cells(f"A{row_idx}:A{row_idx + 1}")
    ws[f"A{row_idx}"] = "No"

    ws.merge_cells(f"B{row_idx}:B{row_idx + 1}")
    ws[f"B{row_idx}"] = "MODE"

    ws.merge_cells(f"C{row_idx}:E{row_idx}")
    ws[f"C{row_idx}"] = "LIMIT"

    ws[f"C{row_idx + 1}"] = "MIN."
    ws[f"D{row_idx + 1}"] = "Typ"
    ws[f"E{row_idx + 1}"] = "MAX."

    ws.merge_cells(f"F{row_idx}:F{row_idx + 1}")
    ws[f"F{row_idx}"] = "UNIT"

    ws.merge_cells(f"G{row_idx}:G{row_idx + 1}")
    ws[f"G{row_idx}"] = "CONDITIONS"

    for r in range(row_idx, row_idx + 2):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

    row_idx += 2
    # Placeholder for characteristics loop
    # We will handle dynamic rows in the actual export logic,
    # but for the template we can leave a few blank rows or a marker
    ws[f"A{row_idx}"] = "{{characteristics_start}}"

    # Notes
    row_idx += 10  # Leave space for characteristics
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    ws[f"A{row_idx}"] = (
        "The F/T specifications must be relaxed than the wafer probing specifications"
    )

    row_idx += 1
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    ws[f"A{row_idx}"] = "*{{esd_display}}"

    row_idx += 2
    notes = [
        ("Sample probe", "400dice or over / wafer for all wafers."),
        (
            "Probing",
            "100% probing is not allowed. The wafers shall be sample probed in order to guarantee the minimum yield.",
        ),
        (
            "Inking",
            "Marking will be made for those areas of wafer cramp nail and process control monitor.",
        ),
        ("Ink type", "Black ink of Epoxy type."),
    ]

    for label, content in notes:
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws[f"A{row_idx}"] = label
        ws[f"A{row_idx}"].font = bold_font

        ws.merge_cells(f"C{row_idx}:G{row_idx}")
        ws[f"C{row_idx}"] = content
        ws[f"C{row_idx}"].alignment = Alignment(wrap_text=True)

        row_idx += 1

    # Footer Note Table
    row_idx += 2
    ws[f"A{row_idx}"] = "NOTE:"
    ws[f"A{row_idx}"].font = bold_font

    row_idx += 1
    headers = ["TYPE: {{sheet_name}}", "TOP METAL", "CHIP THICKNESS", "BACK METAL"]

    ws.merge_cells(f"A{row_idx}:B{row_idx}")
    ws[f"A{row_idx}"] = headers[0]
    ws[f"C{row_idx}"] = headers[1]
    ws.merge_cells(f"D{row_idx}:E{row_idx}")
    ws[f"D{row_idx}"] = headers[2]
    ws.merge_cells(f"F{row_idx}:G{row_idx}")
    ws[f"F{row_idx}"] = headers[3]

    for col in ["A", "C", "D", "F"]:
        cell = ws[f"{col}{row_idx}"]
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    row_idx += 1
    ws[f"A{row_idx}"] = "{{related_devices_start}}"

    # Save
    output_dir = "backend/data"
    os.makedirs(output_dir, exist_ok=True)
    wb.save(os.path.join(output_dir, "spec_sheet_template.xlsx"))
    print(f"Template saved to {os.path.join(output_dir, 'spec_sheet_template.xlsx')}")


if __name__ == "__main__":
    create_template()
